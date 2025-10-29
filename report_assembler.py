import pandas as pd
import numpy as np
import os
import timeit
import win32com.client
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill
from copy import copy
import datetime
import calendar
import locale


locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8') 

today = datetime.datetime.today().strftime('%d.%m.%Y')

# Вводные данные
input_year = 2025
input_month = 3
pomesiachniy = 80
pogodovoi = 20

pd.options.mode.chained_assignment = None  # default='warn'

# Таймер
start = timeit.default_timer()

# Указание путей к файлам с данными
p1 = os.path.abspath('30.04_auto_load\Выгрузка_в_xls_данных_по_запросам_№2194_24_04_2025_11_21.xlsx')
data = pd.read_excel(p1)
p2 = os.path.abspath('30.04_auto_load\файл_показатели_СВПО_14_04_2025_СМЭВ_3.xlsx')
indicators = pd.read_excel(p2)
p_podtbl = os.path.abspath('30.04_auto_load\подтаблица_показатели_СВПО.xlsx')
podtable = pd.read_excel(p_podtbl)

# Загрузка шапки для УДП
p3 = os.path.abspath('30.04_auto_load\Отчет_по_автоинцидентам_УДП.xlsx_15264-Вх_19_03_2025(ver1).xlsx')
df_udp = pd.read_excel(p3,
                       sheet_name='Сведения о НП',
                       dtype=object
                      )

# Для диаграмм
p_diag = os.path.abspath('30.04_auto_load\Для диаграмм.xlsx')

np_map = {
}
np_needed_map = {
}
foiv_map = {
}
foiv_map_reversed = {v: k for k, v in foiv_map.items()}

list_of_svpo = {
}


# Загрузка данных из Выгрузки в УДП
df_udp["НОМЕР ЗАПРОСА"] = data["Номер запроса"].values
df_udp["ДАТА РЕГИСТРАЦИИ"] = data["Дата регистрации"].values
df_udp["ВИД ЗАПРОСА"] = data["Вид запроса"].values
df_udp["СТАТУС"] = data["Системный статус"].values
df_udp["ФЕДЕРАЛЬНЫЙ ПРОЕКТ"] = data["Федеральный проект"].values
df_udp["ID \nПОКАЗАТЕЛЯ"] = data["ID показателя / мероприятия (результата)"].values
df_udp["НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"] = data["Показатель"].values
df_udp["ТИП ПОКАЗАТЕЛЯ"] = data["Тип значения показателя"].values
df_udp["ПЛАН НА КОНЕЦ ГОДА"] = data["Плановое значение на год"].values
df_udp["ОТЧЕТНЫЙ ПЕРИОД"] = data["Отчетный период"].values
# df_udp["ОТЧЕТНЫЙ ПЕРИОД"] = (pd.to_datetime(data["Отчетный период"], format="%d.%m.%Y").dt.strftime("%Y, %B").replace(month_map, regex=True))
df_udp["ПЛАН"] = data["План на отчетный период"].values
df_udp["ФАКТ"] = data["Факт за отчетный период"].values
df_udp["УДП"] = data["Уровень достижения"].values
df_udp["КЛАССИФИКАЦИЯ ИНЦИДЕНТА"] = data["Тип запроса"].values
df_udp["ПОВТОРЯЕМОСТЬ"] = data["Повторяемость"].values
df_udp["ДЛИТЕЛЬНОСТЬ"] = data["Длительность"].values
df_udp["ДАТА ЗАКРЫТИЯ ИНЦИДЕНТА"] = (pd.to_datetime(data["Фактическое время выполнения"], format="%d.%m.%Y")).dt.strftime('%d.%m.%Y')
df_udp["ОПИСАНИЕ РЕШЕНИЯ"] = data["Результат работ"].values
df_udp["ОТВЕТСТВЕННЫЙ ФОИВ"] = data["Ответственный ФОИВ"].values
df_udp["НАЦИОНАЛЬНЫЕ ПРОЕКТЫ"] = data["Национальный проект"].values
df_udp['Результат работ'] = data['Результат работ'].values

df_udp_extra = df_udp[["НОМЕР ЗАПРОСА", "ДАТА РЕГИСТРАЦИИ","ВИД ЗАПРОСА",'ID \nПОКАЗАТЕЛЯ',"НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ","ОТЧЕТНЫЙ ПЕРИОД","ОТВЕТСТВЕННЫЙ ФОИВ","НАЦИОНАЛЬНЫЕ ПРОЕКТЫ"]]
df_udp_extra['Дефицит оборудования и ПО'] = data['Дефицит оборудования и ПО'].values
df_udp_extra['Дефицит человеческих ресурсов'] = data['Дефицит человеческих ресурсов'].values
df_udp_extra['Изменение силы воздействия показателей факторов внешней среды  и их перечня'] = data['Изменение силы воздействия показателей факторов внешней среды  и их перечня'].values
df_udp_extra['Недостатки нормативно-правовых актов'] = data['Недостатки нормативно-правовых актов'].values
df_udp_extra['Недостаточно полномочий сотрудников и подразделений'] = data['Недостаточно полномочий сотрудников и подразделений'].values
df_udp_extra['Недофинансиров..'] = data['Недофинансиров..'].values
df_udp_extra['Не исполнение контрактных обязательств подрядчиками'] = data['Не исполнение контрактных обязательств подрядчиками'].values
df_udp_extra['Некорректная работа методик, алгоритмов'] = data['Некорректная работа методик, алгоритмов'].values
df_udp_extra['Некорректное функционирование информационных систем'] = data['Некорректное функционирование информационных систем'].values
df_udp_extra['Несвоевременное внесение данных в информационные системы'] = data['Несвоевременное внесение данных в информационные системы'].values
df_udp_extra['Неэффективные межведомственные коммуникации'] = data['Неэффективные межведомственные коммуникации'].values
df_udp_extra['Организация работы ПО по проекту'] = data['Организация работы ПО по проекту'].values
df_udp_extra['Отсутствие необходимых методик, алгоритмов'] = data['Отсутствие необходимых методик, алгоритмов'].values

# Фильтрация по месяцу, году и виду запроса

start_year = input_year
start_month = input_month + 1
start_day = 17
end_year = input_year
end_month = input_month + 2
end_day = 16
if start_month > 12:
    start_month = start_month - 12
    start_year += 1
if end_month > 12:
    end_month = end_month - 12
    end_year += 1
list_of_incidents = ['Инцидент УДП','Инцидент данных УДП', 'Предупреждение УДП']
list_of_work_result = ['004 - Параметр не достигнут (нарастающий итог)', '05 - Параметр достигнут в следующий плановый период', '08 - Параметр не достигнут (раздельный учет)']
list_of_status = ['В ожидании взятия в работу', 'В работе']
df_filtered = df_udp
df_filtered = df_filtered[((df_filtered['ВИД ЗАПРОСА'] == 'Инцидент УДП') & 
                                  ((pd.to_datetime(df_filtered['ОТЧЕТНЫЙ ПЕРИОД'], format='%d.%m.%Y').dt.year == input_year) &
                                    (pd.to_datetime(df_filtered['ОТЧЕТНЫЙ ПЕРИОД'], format='%d.%m.%Y').dt.month == input_month))) |
                                        ((df_filtered['ВИД ЗАПРОСА'] == 'Инцидент данных УДП') &
                                        (((df_filtered['ДАТА РЕГИСТРАЦИИ'].dt.year == start_year) & (df_filtered['ДАТА РЕГИСТРАЦИИ'].dt.month == start_month) & (df_filtered['ДАТА РЕГИСТРАЦИИ'].dt.day >= start_day)) |
                                        ((df_filtered['ДАТА РЕГИСТРАЦИИ'].dt.year == end_year) & (df_filtered['ДАТА РЕГИСТРАЦИИ'].dt.month == end_month) & (df_filtered['ДАТА РЕГИСТРАЦИИ'].dt.day <= end_day))) &
                                        (df_filtered['СТАТУС'].isin(list_of_status)) &
                                        (df_filtered['Результат работ'].isin(list_of_work_result))) |
                                        ((df_filtered['ВИД ЗАПРОСА'] == 'Предупреждение УДП') & 
                                        ((pd.to_datetime(df_filtered['ОТЧЕТНЫЙ ПЕРИОД'], format='%d.%m.%Y').dt.year == input_year) &
                                        (pd.to_datetime(df_filtered['ОТЧЕТНЫЙ ПЕРИОД'], format='%d.%m.%Y').dt.month == input_month)))
                                        
                                        ]

# Для таблицы 3 Раздела 1
spravka_date_end = '16.02.' + str(input_year+1)
spravka_date_start = '17.02.' + str(input_year)
year_diff = (pd.to_datetime(spravka_date_end, dayfirst = True) - pd.to_datetime(spravka_date_start, dayfirst = True)).days
df_filtered1 = df_udp
df_for_spr = df_filtered1[((df_filtered1['ВИД ЗАПРОСА'] == 'Инцидент УДП') & 
                                  (pd.to_datetime(df_filtered1['ОТЧЕТНЫЙ ПЕРИОД'], format='%d.%m.%Y').dt.year == input_year)) |
                                        ((df_filtered1['ВИД ЗАПРОСА'] == 'Инцидент данных УДП') &
                                        ((0 <= (pd.to_datetime(spravka_date_end, dayfirst = True) - 
                                               pd.to_datetime(df_filtered1['ДАТА РЕГИСТРАЦИИ'], format='%d.%m.%Y')).dt.days) &
                                         ((pd.to_datetime(spravka_date_end, dayfirst = True) - 
                                               pd.to_datetime(df_filtered1['ДАТА РЕГИСТРАЦИИ'], format='%d.%m.%Y')).dt.days <= year_diff))) 
                                        |
                                        ((df_filtered1['ВИД ЗАПРОСА'] == 'Предупреждение УДП') & 
                                        (pd.to_datetime(df_filtered1['ОТЧЕТНЫЙ ПЕРИОД'], format='%d.%m.%Y').dt.year == input_year))                                        
                                        ]

# Список ФОИВ для Таблицы 2
df_for_name_foiv = df_udp[df_udp['ДАТА РЕГИСТРАЦИИ'].dt.year == 2024]
np_filtered = df_udp[
    # (df_filtered['Результат работ'].isin(list_of_work_result)) &
    (df_udp['ВИД ЗАПРОСА'].isin(list_of_incidents))
]

# Создание Таблицы 2 раздела Справка
spravka_fake_header = (
    "НОМЕР",
    "НАЦИОНАЛЬНЫЕ ПРОЕКТЫ",
    "ВПО, ОТВЕТСТВЕННОЕ ЗА НП",
    "КОЛ-ВО ПОКАЗАТЕЛЕЙ",
    "КОЛИЧЕСТВО НЕДОСТИГНУТЫХ ПОКАЗАТЕЛЕЙ",
    "УРОВЕНЬ ДОСТИЖЕНИЯ < 100%",
    "ДАННЫЕ ОТСУТСТВУЮТ",
    "ОШИБКА ПЛАНИРОВАНИЯ",
    "ПОКАЗАТЕЛЬ ДОСТИГНУТ (УТОЧНЕНИЕ ДАННЫХ)"
)

df_spravka_fake = pd.DataFrame(columns=spravka_fake_header)
df_spravka_fake = df_spravka_fake.set_index('НОМЕР')

# Атрибуты: Национальные проекты; ВПО, ответственное за НП; кол-во показателей
indicators_count = indicators[['np_short_name', 'fp_purpouse_criteria_id']].groupby('np_short_name').count().reset_index()
indicators_count['np_short_name'] = indicators_count['np_short_name'].str.replace('"', '')
indicators_count = indicators_count.sort_values(by='np_short_name')

# Короткое название НП
df_spravka_fake['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'] = indicators_count['np_short_name'].values

# Количество показателей НП, за достижение которых ВПО несет ответственность
df_spravka_fake['КОЛ-ВО ПОКАЗАТЕЛЕЙ'] = indicators_count['fp_purpouse_criteria_id'].values



otv_foiv = indicators[['np_short_name', 'np_foiv','ogrn_fp']].groupby(['np_short_name','np_foiv']).count().reset_index()
otv_foiv['np_foiv'] = otv_foiv['np_foiv'].replace(foiv_map_reversed, regex=True)
otv_foiv = otv_foiv.loc[otv_foiv.groupby(['np_short_name'])['ogrn_fp'].idxmax()]
otv_foiv = otv_foiv[['np_short_name', 'np_foiv']]
otv_foiv = otv_foiv.rename(columns={"np_short_name": "НАЦИОНАЛЬНЫЕ ПРОЕКТЫ", "np_foiv": "ОТВЕТСТВЕННЫЙ ФОИВ"})

for i in range(len(otv_foiv['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'])):
    temp = otv_foiv.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ']
    if temp[0:2] == 'НП':
        temp = temp[3:]
    if temp[0] == '"':
        temp = temp[1:-1]
    otv_foiv.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'] = temp

df_spravka_fake = df_spravka_fake.merge(otv_foiv, how='left', on='НАЦИОНАЛЬНЫЕ ПРОЕКТЫ')


# Инциденты в Таблице 2
grouped_indx = df_filtered[['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', 'ВИД ЗАПРОСА']]
grouped_incindent = grouped_indx.groupby(['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', 'ВИД ЗАПРОСА'])['ВИД ЗАПРОСА'].count().unstack()
if 'Предупреждение УДП' not in grouped_incindent:
    grouped_incindent['Предупреждение УДП'] = 0
if 'Инцидент УДП' not in grouped_incindent:
    grouped_incindent['Инцидент УДП'] = 0
if 'Инцидент данных УДП	' not in grouped_incindent:
    grouped_incindent['Инцидент данных УДП'] = 0
grouped_incindent[np.isnan(grouped_incindent)] = 0
grouped_incindent = grouped_incindent.astype('int32')
grouped_incindent = grouped_incindent.reset_index()

for i in range(len(grouped_incindent['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'])):
    temp = grouped_incindent.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ']
    if temp[0:2] == 'НП':
        temp = temp[3:]
    if temp[0] == '"':
        temp = temp[1:-1]
    grouped_incindent.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'] = temp

df_spravka_fake = df_spravka_fake.merge(grouped_incindent, how='left', on='НАЦИОНАЛЬНЫЕ ПРОЕКТЫ')
# смена типа данных
df_spravka_fake['Инцидент УДП'] = df_spravka_fake['Инцидент УДП'].fillna(0)
df_spravka_fake = df_spravka_fake.astype({'Инцидент УДП': 'int32'})
df_spravka_fake['Инцидент данных УДП'] = df_spravka_fake['Инцидент данных УДП'].fillna(0)
df_spravka_fake = df_spravka_fake.astype({'Инцидент данных УДП': 'int32'})
df_spravka_fake['Предупреждение УДП'] = df_spravka_fake['Предупреждение УДП'].fillna(0)
df_spravka_fake = df_spravka_fake.astype({'Предупреждение УДП': 'int32'})
df_spravka_fake['КОЛИЧЕСТВО НЕДОСТИГНУТЫХ ПОКАЗАТЕЛЕЙ'] = df_spravka_fake['Инцидент УДП'] + df_spravka_fake['Инцидент данных УДП'] + df_spravka_fake['Предупреждение УДП']

# Колонка "Показатель достигнут" в Таблице 2
df_pokaz_dostignut = df_udp[(((df_udp['ВИД ЗАПРОСА'] == 'Инцидент УДП') & 
                                  ((pd.to_datetime(df_udp['ОТЧЕТНЫЙ ПЕРИОД'], format='%d.%m.%Y').dt.year == input_year) &
                                    (pd.to_datetime(df_udp['ОТЧЕТНЫЙ ПЕРИОД'], format='%d.%m.%Y').dt.month == input_month))) |

                                  ((df_udp['ВИД ЗАПРОСА'] == 'Инцидент данных УДП') &
                                        (((df_udp['ДАТА РЕГИСТРАЦИИ'].dt.year == start_year) & (df_udp['ДАТА РЕГИСТРАЦИИ'].dt.month == start_month) & (df_udp['ДАТА РЕГИСТРАЦИИ'].dt.day >= start_day)) |
                                        ((df_udp['ДАТА РЕГИСТРАЦИИ'].dt.year == end_year) & (df_udp['ДАТА РЕГИСТРАЦИИ'].dt.month == end_month) & (df_udp['ДАТА РЕГИСТРАЦИИ'].dt.day <= end_day)))))
                                    &
                                        (df_udp['Результат работ'] == '007 - Уточнение данных, параметр достигнут в плановый период')]

# df_pokaz_dostignut = df_filtered
df_pokaz = df_pokaz_dostignut[['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', 'ВИД ЗАПРОСА']].groupby('НАЦИОНАЛЬНЫЕ ПРОЕКТЫ').count()
df_pokaz = df_pokaz.reset_index()
df_pokaz = df_pokaz.rename(columns={"ВИД ЗАПРОСА": "ПОКАЗАТЕЛЬ ДОСТИГНУТ"})
for i in range(len(df_pokaz['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'])):
    temp = df_pokaz.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ']
    if temp[0:2] == 'НП':
        temp = temp[3:]
    if temp[0] == '"':
        temp = temp[1:-1]
    df_pokaz.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'] = temp

df_spravka_fake = df_spravka_fake.merge(df_pokaz, how='left', on='НАЦИОНАЛЬНЫЕ ПРОЕКТЫ')
# Смена данных колонки
df_spravka_fake['ПОКАЗАТЕЛЬ ДОСТИГНУТ'] = df_spravka_fake['ПОКАЗАТЕЛЬ ДОСТИГНУТ'].fillna(0)
df_spravka_fake = df_spravka_fake.astype({'ПОКАЗАТЕЛЬ ДОСТИГНУТ': 'int32'})


# Итоговая таблица 2
df_spravka_table2 = df_spravka_fake[['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ','ОТВЕТСТВЕННЫЙ ФОИВ','КОЛ-ВО ПОКАЗАТЕЛЕЙ','КОЛИЧЕСТВО НЕДОСТИГНУТЫХ ПОКАЗАТЕЛЕЙ','Инцидент УДП','Инцидент данных УДП','Предупреждение УДП','ПОКАЗАТЕЛЬ ДОСТИГНУТ']]

df_spravka_table2.index = df_spravka_table2.index + 1
# Конец  Таблицы 2 Раздела 1

# Раздел 2. ВПО
vpo1 = indicators[['np_short_name','np_foiv','fp_purpouse_criteria_id']]
vpo1 = vpo1.rename(columns={"np_short_name": "НАЦИОНАЛЬНЫЕ ПРОЕКТЫ", "np_foiv": "ВПО"})
for i in range(len(vpo1['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'])):
    temp = vpo1.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ']
    if temp[0:2] == 'НП':
        temp = temp[3:]
    if temp[0] == '"':
        temp = temp[1:-1]
    vpo1.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'] = temp
vpo1 = vpo1.groupby(by = ['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ','ВПО']).count().reset_index()
vpo1 = vpo1.merge(df_spravka_table2, how='left', on='НАЦИОНАЛЬНЫЕ ПРОЕКТЫ')
vpo1 = vpo1.replace(np_map, regex=True)
for i in range(len(vpo1['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'])):
    temp = vpo1.loc[i,'ОТВЕТСТВЕННЫЙ ФОИВ']
    temp1 = vpo1.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ']
    vpo1.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'] = temp1+' \n(' + temp + ')'
vpo1_clean = vpo1[['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', 'ВПО', 'fp_purpouse_criteria_id']]


vnp4 = df_filtered[((df_filtered['ВИД ЗАПРОСА'] == 'Инцидент УДП') | (df_filtered['ВИД ЗАПРОСА'] == 'Инцидент данных УДП')) & (df_filtered['СТАТУС'].isin(list_of_status) | df_filtered['Результат работ'].isin(list_of_work_result)) | (df_filtered['ВИД ЗАПРОСА'] =='Предупреждение УДП')]

vpn5 = vnp4[['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ','ОТВЕТСТВЕННЫЙ ФОИВ','ВИД ЗАПРОСА']].groupby(by = ['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ','ОТВЕТСТВЕННЫЙ ФОИВ']).count().reset_index()
# vpn5 = vpn5.drop(columns=['ОТВЕТСТВЕННЫЙ ФОИВ'])
vpn5 = vpn5.rename(columns={"ОТВЕТСТВЕННЫЙ ФОИВ": "ВПО"})
for i in range(len(vpn5['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'])):
    temp = vpn5.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ']
    if temp[0:2] == 'НП':
        temp = temp[3:]
    if temp[0] == '"':
        temp = temp[1:-1]
    vpn5.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'] = temp
vpn5 = vpn5.merge(df_spravka_table2, how='left', on='НАЦИОНАЛЬНЫЕ ПРОЕКТЫ')
# print(vpn5)
# vpn5 = vpn5.replace(np_map, regex=True)
for i in range(len(vpn5['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'])):
    temp = vpn5.loc[i,'ОТВЕТСТВЕННЫЙ ФОИВ']
    temp1 = vpn5.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ']
    vpn5.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'] = str(temp1)+' \n(' + str(temp) + ')'
vpn5['ВПО'] = vpn5['ВПО'].replace(foiv_map, regex=True)
vpo2_clean = vpn5[['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', 'ВПО', 'ВИД ЗАПРОСА']]
vpo1_clean = vpo1_clean.rename(columns={"fp_purpouse_criteria_id": "ВСЕГО ПОКАЗАТЕЛЕЙ"})
vpo2_clean = vpo2_clean.rename(columns={"ВИД ЗАПРОСА": "ИЗ НИХ ИНЦИНДЕНТ"})
df = pd.merge(vpo1_clean, vpo2_clean, on=['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', 'ВПО'], how = 'outer')
df = df.drop_duplicates()
tbl_vpo_length = len(df['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'].unique())
pivot = df.pivot(index='ВПО', columns='НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', values=['ВСЕГО ПОКАЗАТЕЛЕЙ', 'ИЗ НИХ ИНЦИНДЕНТ'])
pivot = pivot.swaplevel(axis = 1)
pivot = pivot.sort_index(axis=1, level=0)
pivot = pivot.fillna(0)
 
# Исправление НП и ВПО атрибутов в таблице 2 раздела 1
df_spravka_fake['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'] = df_spravka_fake['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'].replace(np_map, regex=True)
for i in range(len(df_spravka_fake['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'])):
    temp = df_spravka_fake.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ']
    if temp[0:2] == 'НП':
        temp = temp[3:]
    if temp[0] == '"':
        temp = temp[1:-1]
    df_spravka_fake.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'] = temp

for i in range(len(df_spravka_fake['ОТВЕТСТВЕННЫЙ ФОИВ'])):
    temp = df_spravka_fake.loc[i,'ОТВЕТСТВЕННЫЙ ФОИВ']
    if temp[0:3] == 'ВПО':
        temp = temp[4:]
    if temp[0] == '"':
        temp = temp[1:-1]
    df_spravka_fake.loc[i,'ОТВЕТСТВЕННЫЙ ФОИВ'] = temp

# Итоговая таблица 2
df_spravka_table2 = df_spravka_fake[['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ','ОТВЕТСТВЕННЫЙ ФОИВ','КОЛ-ВО ПОКАЗАТЕЛЕЙ','КОЛИЧЕСТВО НЕДОСТИГНУТЫХ ПОКАЗАТЕЛЕЙ','Инцидент УДП','Инцидент данных УДП','Предупреждение УДП','ПОКАЗАТЕЛЬ ДОСТИГНУТ']]
# print(df_spravka_table2['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'].unique())
# Поле индекса № 2 приводится к правильному виду
df_spravka_table2.index = df_spravka_table2.index + 1


# Таблица 3 Раздела 1
prich_filtered = df_for_spr[['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', 'ID \nПОКАЗАТЕЛЯ', 'НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ']]

# Лист причин инцидентов
df_udp_extra = df_udp_extra.fillna('')
df_udp_extra['Причины'] = df_udp_extra['Дефицит оборудования и ПО'] + df_udp_extra['Дефицит человеческих ресурсов'] + df_udp_extra['Изменение силы воздействия показателей факторов внешней среды  и их перечня'] + df_udp_extra['Недостатки нормативно-правовых актов'] + df_udp_extra['Недостаточно полномочий сотрудников и подразделений'] + df_udp_extra['Недофинансиров..'] + df_udp_extra['Не исполнение контрактных обязательств подрядчиками'] + df_udp_extra['Некорректная работа методик, алгоритмов'] + df_udp_extra['Некорректное функционирование информационных систем'] + df_udp_extra['Несвоевременное внесение данных в информационные системы'] + df_udp_extra['Неэффективные межведомственные коммуникации'] + df_udp_extra['Организация работы ПО по проекту'] + df_udp_extra['Отсутствие необходимых методик, алгоритмов']
df_reasons= df_udp_extra[((df_udp_extra['ВИД ЗАПРОСА'] == 'Инцидент УДП') & 
                                  (pd.to_datetime(df_udp_extra['ОТЧЕТНЫЙ ПЕРИОД'], format='%d.%m.%Y').dt.year == input_year)) |
                                        ((df_udp_extra['ВИД ЗАПРОСА'] == 'Инцидент данных УДП') &
                                        ((0 <= (pd.to_datetime(spravka_date_end, dayfirst = True) - 
                                               pd.to_datetime(df_udp_extra['ДАТА РЕГИСТРАЦИИ'], format='%d.%m.%Y')).dt.days) &
                                         ((pd.to_datetime(spravka_date_end, dayfirst = True) - 
                                               pd.to_datetime(df_udp_extra['ДАТА РЕГИСТРАЦИИ'], format='%d.%m.%Y')).dt.days <= year_diff))) 
                                        |
                                        ((df_udp_extra['ВИД ЗАПРОСА'] == 'Предупреждение УДП') & 
                                        (pd.to_datetime(df_udp_extra['ОТЧЕТНЫЙ ПЕРИОД'], format='%d.%m.%Y').dt.year == input_year))                                        
                                        ]
df_klass = df_udp_extra[['НОМЕР ЗАПРОСА', 'Причины']]
df_prichini = df_reasons['Причины'].unique()
df_reasons_try =  df_reasons[['НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ','Причины']]
pivot_reasons = pd.crosstab(index=df_reasons_try['НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ'], columns=df_reasons_try['Причины']).reset_index()
prich_filtered = prich_filtered.merge(pivot_reasons, how = 'left', on = 'НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ')
prich_filtered = prich_filtered.drop_duplicates()





# Добавление атрибутов Уточнение данных, Ошибка планирования и Параметр исключен к Таблице 3 Раздела 1
prich_utoch = df_pokaz_dostignut[['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', "НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ", 'ВИД ЗАПРОСА']].groupby(['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', "НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"]).count().reset_index()
prich_utoch = prich_utoch.rename(columns={'ВИД ЗАПРОСА' : 'Уточнение данных, показатель достигнут'})
prich_filtered = prich_filtered.merge(prich_utoch, how = 'left', on = ['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', 'НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ'])
prich_oshib = df_udp[((df_udp['ВИД ЗАПРОСА'] == 'Предупреждение УДП') & 
                                        ((pd.to_datetime(df_udp['ОТЧЕТНЫЙ ПЕРИОД'], format='%d.%m.%Y').dt.year == input_year) &
                                        (pd.to_datetime(df_udp['ОТЧЕТНЫЙ ПЕРИОД'], format='%d.%m.%Y').dt.month == input_month)))]
prich_oshib = prich_oshib[['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', "НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ", 'ВИД ЗАПРОСА']].groupby(['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', "НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"]).count().reset_index()
prich_oshib = prich_oshib.rename(columns={'ВИД ЗАПРОСА' : 'Ошибка планирования'})       
prich_filtered = prich_filtered.merge(prich_oshib, how = 'left', on = ['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', 'НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ'])                 
prich_isk = df_for_spr[df_for_spr['Результат работ'] == '03 Параметр исключен']
prich_isk = prich_isk[['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', "НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ", 'ВИД ЗАПРОСА']].groupby(['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', "НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"]).count().reset_index()
prich_isk = prich_isk.rename(columns={'ВИД ЗАПРОСА' : 'Параметр исключен'})
prich_filtered = prich_filtered.merge(prich_isk, how = 'left', on = ['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', 'НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ'])

# Исправление графы ID Таблицы 3 Раздела 1
prich_filtered = prich_filtered.reset_index()
for i in range(len(prich_filtered["НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"])):
    temp = prich_filtered.loc[i,"НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"]
    if pd.isnull(temp):
        flag = 1
    else:
        if temp[10:13] == ' - ':
            temp_name = temp[13:]
            temp_id = temp[0:10]
        else:
            temp = temp.split()
            temp_id = temp[0]
            temp_name = ''.join(temp[2:])
        if temp_name != "":
            while temp_name[0] == '"':
                temp_name = temp_name[1:-1]
        prich_filtered.loc[i,"ID \nПОКАЗАТЕЛЯ"] = temp_id
        prich_filtered.loc[i,"НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"] = temp_name
prich_filtered = prich_filtered.drop(columns=['index'])
prich_filtered = prich_filtered.fillna('')


# Поле НП приводится к правильному виду
for i in range(len(prich_filtered['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'])):
    temp = prich_filtered.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ']
    if temp[0:2] == 'НП':
        temp = temp[3:]
    if temp[0] == '"':
        temp = temp[1:-1]
    prich_filtered.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'] = temp
prich_filtered['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'] = prich_filtered['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'].replace(np_map, regex=True)
for i in range(len(prich_filtered['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'])):
    temp = prich_filtered.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ']
    if temp[0:2] == 'НП':
        temp = temp[3:]
    if temp[0] == '"':
        temp = temp[1:-1]
    prich_filtered.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'] = temp

# Сортировка по НП
prich_filtered = prich_filtered.sort_values(by=['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'], ignore_index = True)

prich_np = prich_filtered[['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', 'НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ']].groupby('НАЦИОНАЛЬНЫЕ ПРОЕКТЫ').count().reset_index()
# Поле индекса приводится к правильному виду
prich_filtered.index = prich_filtered.index + 1

# Кусок кода на 200 строк чисто для одной функции - покраса ячейки Таблицы 3...
nujno_li_format_tabl_3 = 1
if nujno_li_format_tabl_3 == 1:
    df_reasons1 = df_reasons[((df_reasons['ВИД ЗАПРОСА'] == 'Инцидент УДП') & 
                                  (pd.to_datetime(df_reasons['ОТЧЕТНЫЙ ПЕРИОД'], format='%d.%m.%Y').dt.month == input_month)) |
                                        ((df_reasons['ВИД ЗАПРОСА'] == 'Инцидент данных УДП') &
                                        (((df_reasons['ДАТА РЕГИСТРАЦИИ'].dt.year == start_year) & (df_reasons['ДАТА РЕГИСТРАЦИИ'].dt.month == start_month) & (df_reasons['ДАТА РЕГИСТРАЦИИ'].dt.day >= start_day)) |
                                        ((df_reasons['ДАТА РЕГИСТРАЦИИ'].dt.year == end_year) & (df_reasons['ДАТА РЕГИСТРАЦИИ'].dt.month == end_month) & (df_reasons['ДАТА РЕГИСТРАЦИИ'].dt.day <= end_day))) &
                                        ((df_reasons['ВИД ЗАПРОСА'] == 'Предупреждение УДП') & 
                                        (pd.to_datetime(df_reasons['ОТЧЕТНЫЙ ПЕРИОД'], format='%d.%m.%Y').dt.month == input_month)))
                                        ]
    
    
    df_prichini1 = df_reasons1['Причины'].unique()
    df_reasons_try1 =  df_reasons1[['НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ','Причины']]
    prich_filtered1 = df_reasons1[['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ', 'ID \nПОКАЗАТЕЛЯ', 'НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ']]
    pivot_reasons1 = pd.crosstab(index=df_reasons_try1['НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ'], columns=df_reasons_try1['Причины']).reset_index()
    prich_filtered1 = prich_filtered1.merge(pivot_reasons1, how = 'left', on = 'НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ')
    prich_filtered1 = prich_filtered1.drop_duplicates()

    # Исправление графы ID Таблицы 3 Раздела 1
    prich_filtered1 = prich_filtered1.reset_index()
    for i in range(len(prich_filtered1["НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"])):
        temp = prich_filtered1.loc[i,"НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"]
        if pd.isnull(temp):
            flag = 1
        else:
            if temp[10:13] == ' - ':
                temp_name = temp[13:]
                temp_id = temp[0:10]
            else:
                temp = temp.split()
                temp_id = temp[0]
                temp_name = ''.join(temp[2:])
            if temp_name != "":
                while temp_name[0] == '"':
                    temp_name = temp_name[1:-1]
            # temp = temp.split()
            # if temp[1] == '-':
            #     temp_id = temp[0]
            #     temp_name = temp[2:].join()
            prich_filtered1.loc[i,"ID \nПОКАЗАТЕЛЯ"] = temp_id
            prich_filtered1.loc[i,"НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"] = temp_name
    prich_filtered1 = prich_filtered1.drop(columns=['index'])
    prich_filtered1 = prich_filtered1.fillna('')

# Раздел 3

# Обновление внешней таблицы подложки
mon1 = podtable.loc[1, 'date_calculation']
mon2 = indicators.loc[1, 'date_calculation']
indicators = indicators.reset_index()
podtable = podtable.reset_index()
if  int(mon1[5:7]) < int(mon2[5:7]):
    value = 'value_' + str(input_month)
    plan = 'plan_' + str(input_month)
    level_ach = 'level_ach_' + str(input_month)
    for i in range(len(indicators)):
        name = indicators.loc[i, 'fp_purpouse_criteria_id']
        for j in range(len(podtable)):
            if podtable.loc[j, 'fp_purpouse_criteria_id'] == name:
                date = indicators.loc[0, 'per_key']
                if int(date[5:]) == input_month:
                    podtable.loc[j, value] = indicators.loc[i, 'value_m']
                    podtable.loc[j, plan] = indicators.loc[i, 'plan_month']
                    podtable.loc[j, level_ach] = indicators.loc[i, 'level_ach']
                    podtable.loc[j,'pc_comment'] = podtable.loc[j,'pc_comment'] + indicators.loc[i,'pc_comment']
                    podtable.loc[j,'date_calculation'] = indicators.loc[i,'date_calculation']
                # else:
                # За старые периоды не должно обновляться, согласно 5.2
    # podtable[value] = indicators['value_m']
    # podtable[plan] = indicators['plan_month']
    # podtable[level_ach] = indicators['level_ach']
    # podtable['date_calculation'] = indicators['date_calculation']
    podtable.to_excel('подтаблица_показатели_СВПО.xlsx', sheet_name='Sheet',merge_cells=True) 
    

# (pd.to_datetime(spravka_date_end, dayfirst = True) - pd.to_datetime(df_filtered1['ДАТА РЕГИСТРАЦИИ'], format='%d.%m.%Y')).dt.days)

podlojka_vse = podtable[['fp_short_name', 'fp_foiv' , 'fp_short_name', 'fp_purpouse_criteria_id', 'fp_purpouse_criteria_name' , 'fp_pc_type', 'okei_criteria_name']]
podlojka_vse['vivodi'] = 0
for i in range(input_month):
    value = 'value_' + str(i+1)
    plan = 'plan_' + str(i+1)
    level_ach = 'level_ach_' + str(i+1)
    podlojka_vse[value] = podtable[value]
    podlojka_vse[plan] = podtable[plan]
    podlojka_vse[level_ach] = podtable[level_ach]
    for j in range(len(podlojka_vse)):
        temp = podlojka_vse.loc[j, level_ach]
        if pd.isnull(temp) == False:
            if int(temp) < 100:
                podlojka_vse.loc[j, 'vivodi'] = 1

podlojka = podlojka_vse[podlojka_vse['vivodi'] == 1]
podlojka = podlojka.drop(columns=['vivodi'])

podlojka['fp_foiv'] = podlojka['fp_foiv'].replace(foiv_map_reversed, regex=True)
podlojka = podlojka.reset_index()
podlojka = podlojka.drop(columns=['index'])

for i in range(len(podlojka['fp_foiv'])):
    temp = podlojka.loc[i,'fp_foiv']
    if temp[0:3] == 'ВПО':
        temp = temp[4:]
    podlojka.loc[i,'fp_foiv'] = temp



# Создание внешней таблицы для хистограммы Раздела 3
cols = {'Месяц': ['янв', 'фев', "мар", "апр", "май", "июн", "июл", "авг", "сен", "окт", "ноя", "дек"],
         'не достигнуто показателей': [np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan],
         'данные отсутствуют': [np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan],
         'помесячный': [np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan]
}
df_hist = pd.DataFrame(cols)

for i in range(input_month):
    sum_nedost = 0
    dan_ots = 0
    level_ach = 'level_ach_' + str(i+1)
    for j in range(len(podlojka)):
        if podlojka.loc[j, level_ach] != podlojka.loc[j, level_ach]:
            dan_ots += 1
        else:
            val = int(podlojka.loc[j, level_ach])
        if 0 < val < 100:
            sum_nedost += 1
    df_hist.loc[i, 'не достигнуто показателей'] = sum_nedost
    df_hist.loc[i, 'данные отсутствуют'] = dan_ots
    df_hist.loc[i, 'помесячный'] = len(podlojka)
# df_hist = df_hist.astype('str')
for i in range(input_month, 12):
    df_hist.loc[i, 'не достигнуто показателей'] = 0
    df_hist.loc[i, 'данные отсутствуют'] = 0
    df_hist.loc[i, 'помесячный'] = 0    

df_hist = df_hist.drop(columns=['помесячный'])
df_hist.to_excel(p_diag, sheet_name = 'Диаграмма 2',merge_cells=True, index= False) 

# Запись в итоговый файл

excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False # Change to True if you want to see it
# excel.Visible = True

# Open the workbook
wb = excel.Workbooks.Open(p3)

# r'c:\python\practicing\30.04_auto_load\Отчет_по_автоинцидентам_УДП.xlsx_15264-Вх_19_03_2025(ver1).xlsx'
ws = wb.Sheets("Справка") 

# С годовым и помесячным
ws.Cells(6, 9).Value = 20
ws.Cells(7, 9).Value = 80

# Define where to insert (e.g., insert 3 rows starting from row 5)
insert_at = 18
rows_to_insert = len(df_spravka_table2) - 2

# Insert rows with formatting (like in Excel: Right-click > Insert > Entire row)
for i in range(rows_to_insert):
    ws.Rows(insert_at).Insert(Shift=-4121)  # -4121 is xlShiftDown

    # Copy formatting from the row above
    ws.Rows(insert_at).Copy()
    ws.Rows(insert_at).PasteSpecial(Paste=-4122)  # -4122 is xlPasteFormats

df_spravka_table2 = df_spravka_table2.reset_index()
# Запись данных Таблицы 2 Раздела 1
start_row = 16
start_col = 2
for row_idx, row in df_spravka_table2.iterrows():
    for col_idx, value in enumerate(row):
        ws.Cells(int(start_row) + 1 + int(row_idx), int(start_col) + int(col_idx)).Value = value

# Форматирование данных в Таблице 2
length = len(df_spravka_table2)+17
range_str = "F17:F"+str(length - 1)
target_range = ws.Range(range_str) 
for row in target_range.Rows:
    for cell in row.Cells:
        if cell.Value == 0:
            cell.Font.Color = 0x00FF00
        else:
            cell.Font.Color = 0x00008B

range_str = 'G17:I' + str(length - 1)
target_range = ws.Range(range_str) 
for row in target_range.Rows:
    for cell in row.Cells:
        if cell.Value == 0:
            cell.Font.Color = 0xD3D3D3
        else:
            cell.Font.Color = 0x00008B

range_str = 'J17:J' + str(length - 1)
target_range = ws.Range(range_str) 
for row in target_range.Rows:
    for cell in row.Cells:
        if cell.Value == 0:
            cell.Font.Color = 0xD3D3D3
        else:
            cell.Font.Color = 0x00FF00


# Вставка столбцов в Таблицу 3 Раздела 1
insert_at = 12
cols_to_insert = (len(df_prichini) - 7)
# print(cols_to_insert)
for i in range(cols_to_insert):
    ws.Columns(insert_at).Insert(Shift=-4161)  # -4121 is xlShiftDown -4161 is xlShiftToRight 

    # Copy formatting 
    ws.Columns(insert_at).Copy()
    ws.Columns(insert_at).PasteSpecial(Paste=-4122)  # -4122 is xlPasteFormats

# Удаление столбцов в случае избыточности
if (len(df_prichini) - 7) < 0:
    letter_start = get_column_letter( 5 + len(df_prichini) + 1)
    # print(df_prichini)
    letter_end = 'L'
    number_start = 23 + len(df_spravka_table2) - 2
    number_end = 23 + len(df_spravka_table2) - 2 + 5
    rng = letter_start + str(number_start) + ':' + letter_end + str(number_end)
    ws.Range(rng).ClearContents()
    rng = letter_start + str(number_start) + ':' + get_column_letter( 5 + len(df_prichini) + 5) + str(number_end)
    # print(rng)
    target_rng = 'M' + str(number_start) + ':' + 'Q' + str(number_end)
    ws.Range(target_rng).Copy(ws.Range(rng))
    target_to_erase =  get_column_letter(5 + len(df_prichini) + 1 + 5) + str(number_start) + ':' + 'Q' + str(number_end)
    # ws.Cells(5 + len(df_prichini) + 1, number_start) = ws.Cells(13, number_start)
    ws.Range('L3:L4').Copy(ws.Range(target_to_erase))


# Вставка строк в Таблицу 3 Раздела 1
insert_at = 27 + len(df_spravka_table2) - 2
rows_to_insert = len(prich_filtered) - 2

# Insert rows with formatting (like in Excel: Right-click > Insert > Entire row)
for i in range(rows_to_insert):
    ws.Rows(insert_at).Insert(Shift=-4121)  # -4121 is xlShiftDown

    # Copy formatting from the row above
    ws.Rows(insert_at).Copy()
    ws.Rows(insert_at).PasteSpecial(Paste=-4122)  # -4122 is xlPasteFormats

# Запись данных Таблицы 3 Раздела 1
prich_filtered = prich_filtered.reset_index()
start_row = 26 + len(df_spravka_table2) - 2
start_col = 2
for row_idx, row in prich_filtered.iterrows():
    for col_idx, value in enumerate(row):
        if value != 0:
            ws.Cells(int(start_row) + int(row_idx), int(start_col) + int(col_idx)).Value = value

list_prich = list(prich_filtered.columns)
for i in range(len(df_prichini)):
    ws.Cells(24 + len(df_spravka_table2) - 2, 6 + i).Value = list_prich[i+4]

# Cells merge
row = 26 + len(df_spravka_table2) - 2
end_col = 10 + len(df_prichini)
for j in range(len(prich_np)):
    col_rows = prich_np.loc[j, 'НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ']
    cells = 'C' + str(row) + ':C' + str(row+col_rows-1)
    rng = ws.Range(cells)
    formula = '=' + get_column_letter(end_col - 1) + str(row)
    for i in range(row, row + col_rows):
        ws.Cells(i, 3).Value = ''  

        itog = '=F' + str(i)
        for k in range(1, len(df_prichini)):
            letter = get_column_letter(k+6)
            itog = itog + '+' + letter + str(i)
        ws.Cells(i, end_col - 1).Value = itog
        if i != row:
            formula = formula + '+' + get_column_letter(end_col - 1) + str(i)
    rng.Merge()
    cell = 'C' + str(row)
    ws.Cells(row, 3).Value = prich_np.loc[j, 'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ']

    cells = get_column_letter(end_col) + str(row) + ':' + get_column_letter(end_col) + str(row+col_rows-1)
    rng = ws.Range(cells)
    rng.Merge()
    ws.Cells(row, end_col).Value = formula
    row = row + col_rows

letter = get_column_letter(len(df_prichini) + 5)
num = 23 + len(df_spravka_table2) - 2
cells = 'F' +str(num) + ':' + letter + str(num)
rng = ws.Range(cells)
rng.Merge()


# Изменение строки ИТОГО Таблицы 3 Раздела 1
for i in range(len(df_prichini) + 5):
    ws.Cells(25 + len(df_spravka_table2) - 2, i+5).value = i + 4
    num = len(prich_filtered) + 26 + len(df_spravka_table2) - 2
    txt = '=SUM(' + get_column_letter(i+6) + str(26 + len(df_spravka_table2) - 2) + ':' + get_column_letter(i+6) + str(num - 1) +')'
    ws.Cells(num,i+6).value = txt



ws = wb.Sheets("ВПО") 

# Define where to insert
insert_at = 20
cols_to_insert = (tbl_vpo_length - 8) * 2
# print(cols_to_insert)
for i in range(cols_to_insert):
    ws.Columns(insert_at).Insert(Shift=-4161)  # -4121 is xlShiftDown -4161 is xlShiftToRight 

    # Copy formatting 
    if i % 2 == 0:
        ws.Columns(insert_at).Copy()
        ws.Columns(insert_at).PasteSpecial(Paste=-4122)  # -4122 is xlPasteFormats
    else:
        ws.Columns(insert_at-2).Copy()
        ws.Columns(insert_at).PasteSpecial(Paste=-4122)  # -4122 is xlPasteFormats

# Вставить строки
insert_at = 10
rows_to_insert = len(pivot) - 2

# Insert rows with formatting
for i in range(rows_to_insert):
    ws.Rows(insert_at).Insert(Shift=-4121)  # -4121 is xlShiftDown

    # Copy formatting from the row above
    ws.Rows(insert_at).Copy()
    ws.Rows(insert_at).PasteSpecial(Paste=-4122)  # -4122 is xlPasteFormats

number = (rows_to_insert + 8)
for_t11 = '=D' + str(rows_to_insert + 11) + '+H' + str(rows_to_insert + 11) + '+L' + str(rows_to_insert + 11)+ '+N' + str(rows_to_insert + 11) + '+P' + str(rows_to_insert + 11) + '+F' + str(rows_to_insert + 11) + '+J' + str(rows_to_insert + 11) + '+R' + str(rows_to_insert + 11)
for_u11 = '=E' + str(rows_to_insert + 11) + '+I' + str(rows_to_insert + 11) + '+M' + str(rows_to_insert + 11)+ '+O' + str(rows_to_insert + 11) + '+Q' + str(rows_to_insert + 11) + '+G' + str(rows_to_insert + 11) + '+K' + str(rows_to_insert + 11) + '+S' + str(rows_to_insert + 11)


# Cells merge
for i in range(20, (tbl_vpo_length - 8) * 2 + 20):
    if i % 2 == 0:
        letter1 = get_column_letter(i)
        
        for_t11 = for_t11 + '+' + letter1 + str(rows_to_insert + 11)
        letter2 = get_column_letter(i+1)
        for_u11 = for_u11 + '+' + letter2 + str(rows_to_insert + 11)
        cells = letter1 + '6:' + letter2 + '6'
        rng = ws.Range(cells)
        rng.Merge()

        range_sum = '=SUM(' + letter1 + '9:' + letter1 + str(rows_to_insert + 10) +')'
        ws.Cells((rows_to_insert + 11), i).Value = range_sum
        range_sum = '=SUM(' + letter2 + '9:' + letter2 + str(rows_to_insert + 10) +')'
        ws.Cells((rows_to_insert + 11), i+1).Value = range_sum
        

# Вставка названий НП
for i in range(pivot.shape[1]):
    # print(i)
    if i % 2 == 0:
        a = pivot.columns[i]
        ws.Cells(6, 4 + i).Value = a[0]

# Запись данных в ВПО
if isinstance(df.index, pd.MultiIndex):
    df = df.reset_index()

pivot = pivot.reset_index()
start_row = 8
start_col = 3
for row_idx, row in pivot.iterrows():
    for col_idx, value in enumerate(row):
        if value != 0:
            ws.Cells(int(start_row) + 1 + int(row_idx), int(start_col) + int(col_idx)).Value = value
            # if (start_col + col_idx) % 2 == 1:
            #     cell = ws.Cells(int(start_row) + 1 + int(row_idx), int(start_col) + int(col_idx))
            #     cell.fill = 0x00008B


# Перезапись отдельных ячеек в ВПО
ws.Cells((rows_to_insert + 11), cols_to_insert + 20).Value = for_t11
ws.Cells((rows_to_insert + 11), cols_to_insert + 21).Value = for_u11

for i in range(9, 9 + rows_to_insert + 2):
    insrt1 = '=D'+str(i)
    insrt2 = '=E'+str(i)
    for j in range(tbl_vpo_length * 2):
        letter = get_column_letter(j+4)
        if j % 2 == 0:
            insrt1 = insrt1 + '+' + letter + str(i)
        else:
            insrt2 = insrt2 + '+' + letter + str(i)
    ws.Cells(i, cols_to_insert + 20).Value = insrt1
    ws.Cells(i, cols_to_insert + 21).Value = insrt2

for i in range(tbl_vpo_length + 1):
    txt = 'из них инцидент на ' + today
    ws.Cells(7, 5 + i * 2).Value = txt
    txt1 = 'ВСЕГО ПОКАЗАТЕЛЕЙ'
    ws.Cells(7, 4 + i * 2).Value = txt1

for i in range(20, (tbl_vpo_length) * 2 + 4):
    ws.Cells(8, i).Value = i - 2

# Заполнение Раздела 3
ws = wb.Sheets("Подложка")

# Вставить строки
insert_at = 23
rows_to_insert = len(podlojka) - 2

# Insert rows with formatting
for i in range(rows_to_insert):
    ws.Rows(insert_at).Insert(Shift=-4121)  # -4121 is xlShiftDown

    # Copy formatting from the row above
    ws.Rows(insert_at).Copy()
    ws.Rows(insert_at).PasteSpecial(Paste=-4122)  # -4122 is xlPasteFormats

# Запись данных
podlojka.index = podlojka.index + 1
podlojka = podlojka.reset_index()
start_row = 22
start_col = 2
for row_idx, row in podlojka.iterrows():
    for col_idx, value in enumerate(row):
        if value != 0:
            ws.Cells(int(start_row) + int(row_idx), int(start_col) + int(col_idx)).Value = value



# Диаграммы

data_sheet = wb.Sheets("Справка")
chart_sheet = wb.Sheets("Подложка")

# Access chart
chart_obj = chart_sheet.ChartObjects(1)
chart = chart_obj.Chart
rng = data_sheet.Range("I6:I7")
rng_lbl = data_sheet.Range('C6:C7')
chart.SeriesCollection(1).XValues = rng_lbl
chart.SeriesCollection(1).Values = rng

# === Enable and customize data labels ===
series = chart.SeriesCollection(1)
series.HasDataLabels = True
labels = series.DataLabels()
labels.ShowCategoryName = False
labels.ShowValue = True
labels.ShowPercentage = False  # Optional

def rgb(r, g, b):
    return r + (g << 8) + (b << 16)

point1 = series.Points(2)
point1.Format.Fill.ForeColor.RGB = rgb(0, 102, 204)  # Blue
point1.Format.Fill.Solid()
line = point1.Format.Line
line.Visible = True
line.ForeColor.RGB = rgb(0, 0, 0)  # Black border
line.Weight = 1  # Thickness in points
line.DashStyle = 1  # 1 = Solid

point2 = series.Points(1)
point2.Format.Fill.Patterned(2)  # 1 = xlDiagonalCross (simulate zebra-like)
line = point2.Format.Line
line.Visible = True
line.ForeColor.RGB = rgb(0, 0, 0)  # Black border
line.Weight = 1  # Thickness in points
line.DashStyle = 1  # 1 = Solid

# point2.Format.Fill.Patterned(5)
point2.Format.Fill.ForeColor.RGB = rgb(0, 0, 0)      # Black
point2.Format.Fill.BackColor.RGB = rgb(173, 216, 230)  # Blue

chart.SeriesCollection(1).Points(1).Explosion = 14


# Способ через вставку диаграммой без ссылки на диапазон (значения прописываются скриптом)

anchor_cell = ws.Range("F4")
left = anchor_cell.Left - 100
top = anchor_cell.Top
width = 600
height = 200

chart_shape = chart_sheet.Shapes.AddChart2(201, 65,  left, top, width, height)
chart = chart_shape.Chart

data_range_ned = []
data_range_net = []
x_labels  = df_hist['Месяц']
for i in range(input_month):
    data_range_ned.append(int(df_hist.loc[i, 'не достигнуто показателей']))
    data_range_net.append(int(df_hist.loc[i, 'данные отсутствуют']))
for i in range(input_month, 12):
    data_range_ned.append('н/д')
    data_range_net.append('н/д')

# Добавление первой линии на граф
line_series = chart.SeriesCollection().NewSeries()
line_series.Values = data_range_ned
line_series.XValues = x_labels
line_series.Name = 'не достигнуто показателей'
line_series.ChartType = 65
line_series.Format.Line.ForeColor.RGB = rgb(70, 148, 225)
# line_series.Format.Line.ForeColor.RGB = rgb(0, 0, 0)
line_series.MarkerBackgroundColor = rgb(0, 58, 135)  # Fill
for i in range(1, len(data_range_ned) + 1):
    point = line_series.Points(i)
    point.HasDataLabel = True
    point.DataLabel.Text = 'yes'
    # point.Format.Fill.ForeColor.RGB = rgb(0, 0, 250)

point = line_series.Points(input_month)
point.HasDataLabel = True
point.DataLabel.Font.Bold = True

# Костыль на убирание лишней линии
if chart.SeriesCollection().Count > 1:
    chart.SeriesCollection(1).Delete()
    print('Erased')

# Добавление второй линии на граф
line_series = chart.SeriesCollection().NewSeries()
line_series.Values = data_range_net
line_series.XValues = x_labels
line_series.Name = 'данные отсутствуют'
line_series.ChartType = 65
line_series.Format.Line.ForeColor.RGB = rgb(255, 115, 50)
line_series.MarkerBackgroundColor = rgb(180, 41, 0)  # Fill
# line_series.MarkerForegroundColor = rgb(0, 0, 139)    # Border

for i in range(1, len(data_range_net) + 1):
    line_series.Points(i).HasDataLabel = True
    line_series.Points(i).DataLabel.Text = data_range_net[i - 1]
    line_series.Points(i).DataLabel.Position = 0
point = line_series.Points(input_month)
point.HasDataLabel = True
point.DataLabel.Font.Bold = True

max_val = 0
for i in range(input_month):
    if data_range_ned[i] > max_val:
        max_val = data_range_ned[i]
    if data_range_net[i] > max_val:
        max_val = data_range_net[i]

# max_val = max(values)
threshold = max_val + 6
line_vals = [threshold] * len(x_labels)

hline = chart.SeriesCollection().NewSeries()
hline.XValues = x_labels
hline.Values = line_vals
hline.Name = "с помесячным планированием"

hline.ChartType = 65  # xlLine
hline.Format.Line.ForeColor.RGB = rgb(200, 200, 200)
hline.MarkerBackgroundColor = rgb(130, 130, 130)  # Fill

# hline.Format.Line.DashStyle = 2
point = hline.Points(input_month)
point.HasDataLabel = True
point.DataLabel.Font.Bold = True

# исправить
max_value = 255
# === Add labels to each point if needed ===
hline.HasDataLabels = True
for i in range(1, hline.Points().Count + 1):
    label = hline.Points(i).DataLabel
    hline.Points(i).DataLabel.Position = 0
    if i > input_month:
      label.Text = ""
      label.Font.Size = 8
      label.Font.Color = rgb(55, 55, 55)
    else:
      label.Text = f"{max_value}"
      label.Font.Size = 8
      label.Font.Color = rgb(55, 55, 55)

text = "Недостижение или отсутствие данных по показателям в " + str(input_year) + ' году'
chart.HasLegend = True
chart.HasTitle = True
chart.ChartTitle.Text = text
chart.ChartTitle.Font.Bold = True
chart.ChartTitle.Left = 10   # Distance from left edge of chart
chart.ChartTitle.Top = 3
# chart.Axes(2).Visible = False
chart.Axes(2).HasMajorGridlines = False
chart.Axes(2).TickLabels.NumberFormat = "" 
# chart.Axes(2).MajorTickMark = 0  # xlNone
# chart.Axes(2).MinorTickMark = 0
chart.Axes(2).Format.Line.Visible = False
chart.Axes(2).TickLabels.Font.Size = 1


path_exit = '30.04_auto_load\\Test_otchet_'
path_exit = path_exit + today + '.xlsx'
# Save and close
p1_exit = os.path.abspath(path_exit)
wb.SaveAs(p1_exit)
# r'c:\python\practicing\30.04_auto_load\Test_otchet_win32.xlsx'
wb.Close()
excel.Quit()

# Исправление графы Отчетный период
np_filtered["ОТЧЕТНЫЙ ПЕРИОД"] = (pd.to_datetime(np_filtered["ОТЧЕТНЫЙ ПЕРИОД"], format="%d.%m.%Y").dt.strftime("%Y, %B"))

# Исправление графы ID Раздела 4
np_filtered = np_filtered.reset_index()
for i in range(len(np_filtered["НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"])):
    temp = np_filtered.loc[i,"НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"]
    if pd.isnull(temp):
        flag = 1
    else:
        if temp[10:13] == ' - ':
            temp_name = temp[13:]
            temp_id = temp[0:10]
        else:
            temp = temp.split()
            temp_id = temp[0]
            temp_name = ''.join(temp[2:])
        if temp_name != "":
            while temp_name[0] == '"':
                temp_name = temp_name[1:-1]
        # temp = temp.split()
        # if temp[1] == '-':
        #     temp_id = temp[0]
        #     temp_name = temp[2:].join()
        np_filtered.loc[i,"ID \nПОКАЗАТЕЛЯ"] = temp_id
        np_filtered.loc[i,"НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"] = temp_name
np_filtered = np_filtered.drop(columns=['index'])

# Исправление названия НП
for i in range(len(np_filtered['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'])):
    temp = np_filtered.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ']
    if pd.isnull(temp):
        flag = 1
    else:
        if temp[0:2] == 'НП':
            temp = temp[3:]
            if temp[0] == '"':
                temp = np.nan
        else:
            temp = np.nan    
        np_filtered.loc[i,'НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'] = temp
np_filtered = np_filtered.dropna(subset=['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'])
np_filtered = np_filtered.reset_index()
np_filtered = np_filtered.drop(columns=['index'])

# Исправление атрбиута Статус
for i in range(len(np_filtered['СТАТУС'])):
    if np_filtered.loc[i,'СТАТУС'] == 'В ожидании взятия в работу':
        np_filtered.loc[i,'СТАТУС'] = 'В работе'

# Заполнеине Раздела 4
    
wb = openpyxl.load_workbook(p1_exit)
wb.active = wb["Сведения о НП"]
ws = wb.active
source = wb.active

# Список уникальных национальных проектов
unique_projects = np_filtered['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'].sort_values().unique()
project_list = []
for project in unique_projects:
    if project in np_needed_map:
        project_list.append(project)
np_filtered = np_filtered.merge(df_klass, how = 'left', on = 'НОМЕР ЗАПРОСА')
np_filtered['КЛАССИФИКАЦИЯ ИНЦИДЕНТА'] = np_filtered['Причины']
np_filtered = np_filtered.drop(columns=['Причины'])
for project in project_list:
    target = wb.copy_worksheet(source)
    target.title = str(project)[:31]
    project_df = np_filtered[(np_filtered['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'] == project)]
    project_df = project_df.drop(columns=['НАЦИОНАЛЬНЫЕ ПРОЕКТЫ'])
    project_df = project_df.drop(columns=['Результат работ'])    
    rows = dataframe_to_rows(project_df, index=False)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            target.cell(row=r_idx, column=c_idx, value=value)
    c = target['D2']
    count = 0
    for row in target:
        if any(cell.value is not None for cell in row):
            count += 1
    last_row = count
    for i in range(1, 20):
        c = target.cell(row = 2, column = i)
        for row in range(2, last_row + 1):
            cell = target.cell(row = row, column = i)
            cell.font = copy(c.font)
            cell.fill = copy(c.fill)
            cell.alignment = copy(c.alignment)

# Форматирование ВПО
wb.active = wb["ВПО"]
ws = wb.active

for j in range(6, 6 + rows_to_insert + 5):
    c = ws.cell(j, 17)
    for i in range(20, tbl_vpo_length * 2 + 4):
        if i % 2 == 1:
            cell = ws.cell(j, i)
            cell.font = copy(c.font)
            cell.fill = copy(c.fill)
            cell.alignment = copy(c.alignment)
            cell.border = copy(c.border)


red_fill = PatternFill(start_color='FFF4CCCC', end_color='FFF4CCCC', fill_type='solid')
start_row = 9
end_row = rows_to_insert + 2 + 8
start_col = 4
end_col = cols_to_insert + 20
for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
    for cell in row:
        if cell.column % 2 == 1 and cell.value not in (None, ""):
            cell.fill = red_fill



del wb['Сведения о НП']

# Заполение Справки
wb.active = wb["Справка"]
ws = wb.active

cell = ws['E3']
cell.value = today
if input_month == 12:
    cell = ws['I5']
    num = len(df_spravka_table2) + 17
    cell_num = 'E' + str(num)
    cell.value = ws[cell_num].value




today_full = datetime.datetime.today()
ws['I8'].value = len(df_for_spr)
txt = calendar.month_name[input_month] + ' ' + str(input_year)
ws['I9'].value = txt

cell = ws['I10']
num = len(df_spravka_table2) + 17
cell_num = 'F' + str(num)
cell.value = ws[cell_num].value

val = 0

if today_full.month  == input_month + 1:
    if today_full.day >= 17:
        val = today
if today_full.month == 1:
    if input_month == 12:
        if today.day >= 24:
            val = today
if val == 0:
    res = calendar.monthrange(input_year, input_month)
    day = res[1]
    if input_month < 10:
        input_month_str = '0' + str(input_month)
    else:
        input_month_str = str(input_month)
    val = str(day) + '.' + input_month_str + '.' + str(input_year)
ws['I11'].value = val

if today_full.year == input_year:
    if val != 0:
        if today_full.month < 11:
            today_month = '0' + str(today_full.month - 1)
        else:
            today_month = str(today_full.month - 1)
        txt = 'ИНЦИДЕНТЫ 01.' + str(input_year) + '-' + today_month + '.' + str(input_year)
    else:
        if today_full.month < 12:
            today_month = '0' + str(today_full.month - 2)
        else:
            today_month = str(today_full.month - 2)
        txt = 'ИНЦИДЕНТЫ 01.' + str(input_year) + '-' + today_month + '.' + str(input_year)
else:
    txt = 'ИНЦИДЕНТЫ 01.' + str(input_year) + '-' + '12.' + str(input_year)
ws['C8'].value = txt

# Изменение ячейки "Всего за ... месяцев" Таблицы 3 Раздела 1
number = 23 + len(df_spravka_table2) - 2
letter = get_column_letter(16+(len(df_prichini) - 7))
rng = letter + str(number)
if input_month > 4:
    ws[rng] = 'ВСЕГО ЗА ' + str(input_month) + ' МЕСЯЦЕВ'   
elif 5 > input_month > 1:
    ws[rng] = 'ВСЕГО ЗА ' + str(input_month) + ' МЕСЯЦА'      
else:
    ws[rng] = 'ВСЕГО ЗА ' + str(input_month) + ' МЕСЯЦ' 



# Сравнение таблиц для окрашивания
if nujno_li_format_tabl_3 == 1:
    red_fill = PatternFill(start_color='FFF4CCCC', end_color='FFF4CCCC', fill_type='solid')
    column_names = list(prich_filtered.columns)
    column_names1 = list(prich_filtered1.columns)
    row = 0
    idx = 0
    for col_idx in range(4, len(column_names1)):
        c_name = column_names1[col_idx]
        idx = column_names.index(c_name)
        for row_idx in range(len(prich_filtered1)):
            r_name = prich_filtered1.iat[row_idx, 2]
            for i in range(len(prich_filtered)):
                if prich_filtered.iat[i, 2] == r_name:
                    row = i
            if prich_filtered1.iat[row_idx, col_idx] != 0:
                if prich_filtered.iat[row, idx] != 0:
                    cell = ws.cell(25 + len(df_spravka_table2) - 2 + row + 1, 5 + idx)
                    cell.fill = red_fill

# Альтернативный способ
if nujno_li_format_tabl_3 == 1:
    red_fill = PatternFill(start_color='BDD7EEDD', end_color='BDD7EEDD', fill_type='solid')
    df_reasons1 = df_reasons1.reset_index()
    for i in range(len(df_reasons1["НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"])):
        temp = df_reasons1.loc[i,"НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"]
        if pd.isnull(temp):
            flag = 1
        else:
            if temp[10:13] == ' - ':
                temp_name = temp[13:]
                temp_id = temp[0:10]
            else:
                temp = temp.split()
                temp_id = temp[0]
                temp_name = ''.join(temp[2:])
            if temp_name != "":
                while temp_name[0] == '"':
                    temp_name = temp_name[1:-1]
            # temp = temp.split()
            # if temp[1] == '-':
            #     temp_id = temp[0]
            #     temp_name = temp[2:].join()
            df_reasons1.loc[i,"ID \nПОКАЗАТЕЛЯ"] = temp_id
            df_reasons1.loc[i,"НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"] = temp_name
    df_reasons1 = df_reasons1.drop(columns=['index'])
    df_reasons1 = df_reasons1.fillna('')


    column_names = list(prich_filtered.columns)
    for i in range(len(df_reasons1)):
        name = df_reasons1.loc[i, "НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"]
        name_prich = df_reasons1.loc[i, "Причины"]
        for j in range(len(prich_filtered)):
            if prich_filtered.loc[j, "НАИМЕНОВАНИЕ ПОКАЗАТЕЛЯ"] == name:
                idx = column_names.index(name_prich)
                cell = ws.cell(25 + len(df_spravka_table2) - 2 + j + 1, 2 + idx)
                cell.fill  = red_fill
                # cell.style.fill  = red_fill
                # cell.style = cell.style.copy(fill = red_fill)

wb.active = wb['Подложка']
ws = wb.active

red_fill = PatternFill(start_color='FFF4CCCC', end_color='FFF4CCCC', fill_type='solid')
green_fill = PatternFill(start_color='CCFFCCCC', end_color='CCFFCCCC', fill_type='solid')
start_row = 22
end_row = len(podlojka) + 22
start_col = 10
end_col = input_month * 3 + 10
for row in ws.iter_rows(min_row = start_row, max_row = end_row, min_col = start_col, max_col = end_col):
    for cell in row:
        if (cell.column - 9) % 3 == 0 and cell.value not in (None, ""):
            if int(cell.value) >= 100:
                cell.value = 100
                cell.fill = green_fill
            else:
                cell.fill = red_fill


wb.save(p1_exit)

stop = timeit.default_timer()

print('Time: ', stop - start) 

print('OK')